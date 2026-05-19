from pathlib import Path
import sqlite3
from openpyxl import load_workbook

DB_PATH = Path("data/unified_tariffs.db")
XLSX_PATH = Path("rules_docs/inbox/ozon/20260426_141844_marketplace-services-rates-01-04-2026.xlsx")

SOURCE_FILE = XLSX_PATH.name
SOURCE_NOTE = "Ozon marketplace services rates, standard tariff table from 01.04.2026"

SHEET_SCHEMES = {
    "FBY с 1.04.2026": "FBY",
    "FBS с 1.04.2026": "FBS",
    "Экспресс с 1.04.2026": "EXPRESS",
    "DBS с 1.04.2026": "DBS",
}

def norm(s):
    return str(s or "").strip().lower().replace("ё", "е")

def to_percent(value):
    if value is None or value == "":
        return None
    try:
        x = float(value)
    except Exception:
        return None

    # В файле Ozon тарифы вида 0.315 = 31.5%
    if 0 < x <= 1:
        return round(x * 100, 4)

    return round(x, 4)

def deepest_category(values):
    clean = [str(v).strip() for v in values if v is not None and str(v).strip()]
    if not clean:
        return "", ""
    category = clean[0]
    product_type = clean[-1]
    return category, product_type

def category_path(values):
    clean = [str(v).strip() for v in values if v is not None and str(v).strip()]
    return " / ".join(clean)

def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"File not found: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Удаляем только предыдущую загрузку этого стандартного файла, Select не трогаем.
    cur.execute(
        "DELETE FROM clean_commissions WHERE marketplace='ozon' AND source_file=?",
        (SOURCE_FILE,)
    )

    inserted = 0

    for sheet_name, scheme in SHEET_SCHEMES.items():
        if sheet_name not in wb.sheetnames:
            print(f"SKIP missing sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"Import sheet: {sheet_name} -> scheme={scheme}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            levels = row[:7]
            tariff_raw = row[7] if len(row) > 7 else None
            fee_percent = to_percent(tariff_raw)

            if fee_percent is None:
                continue

            category, product_type = deepest_category(levels)
            path = category_path(levels)

            if not product_type:
                continue

            # Чтобы не потерять иерархию Ozon, сохраняем путь в category,
            # а самый нижний уровень — в product_type.
            category_for_db = path or category
            product_type_for_db = product_type

            cur.execute(
                """
                INSERT INTO clean_commissions
                (
                    marketplace,
                    category,
                    product_type,
                    scheme,
                    fee_percent,
                    fee_type,
                    valid_from,
                    source_file,
                    source_note,
                    product_type_norm,
                    category_norm
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "ozon",
                    category_for_db,
                    product_type_for_db,
                    scheme,
                    fee_percent,
                    "marketplace_service_rate",
                    "2026-04-01",
                    SOURCE_FILE,
                    SOURCE_NOTE,
                    norm(product_type_for_db),
                    norm(category_for_db),
                )
            )
            inserted += 1

    # Фиксируем качество источника.
    cur.execute(
        """
        INSERT INTO tariff_source_quality
        (marketplace, source_file, source_note, source_status, source_role, comment)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            "ozon",
            SOURCE_FILE,
            SOURCE_NOTE,
            "usable",
            "standard_marketplace_service_rate",
            "Стандартная таблица тарифов услуг Ozon по схемам FBY/FBS/EXPRESS/DBS с 01.04.2026. Использовать для расчёта вместо Ozon Select."
        )
    )

    conn.commit()
    conn.close()

    print(f"Inserted rows: {inserted}")

if __name__ == "__main__":
    main()
