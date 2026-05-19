import csv
import hashlib
import re
from pathlib import Path
from datetime import datetime
from db import init_db, _execute, _fetch_all

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

BASE_DIR = Path("rules_docs/inbox")
SUPPORTED_EXT = {".xlsx", ".csv", ".pdf", ".txt"}

def ensure_tables():
    _execute("""
    CREATE TABLE IF NOT EXISTS rules_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        marketplace TEXT NOT NULL,
        document_name TEXT,
        section TEXT,
        topic TEXT,
        rule_text TEXT NOT NULL,
        effective_date TEXT,
        source_url TEXT,
        content_hash TEXT UNIQUE,
        loaded_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    _execute("""
    CREATE TABLE IF NOT EXISTS rules_imported_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_path TEXT NOT NULL,
        file_size INTEGER,
        file_mtime TEXT,
        file_fingerprint TEXT UNIQUE,
        rows_imported INTEGER DEFAULT 0,
        imported_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

def file_fingerprint(path: Path):
    st = path.stat()
    base = f"{path.resolve()}|{st.st_size}|{st.st_mtime}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:24]

def already_imported(path: Path):
    fp = file_fingerprint(path)
    rows = _fetch_all("SELECT id FROM rules_imported_files WHERE file_fingerprint = ?", (fp,))
    return bool(rows), fp

def infer_marketplace(path: Path):
    parts = [p.lower() for p in path.parts]
    name = path.name.lower()

    if "ozon" in parts or "ozon" in name or "озон" in name:
        return "ozon"
    if "wildberries" in parts or "wb" in parts or "wildberries" in name or "wb" in name or "вайлдберриз" in name:
        return "wildberries"
    if "yandex_market" in parts or "yandex" in name or "яндекс" in name or "market" in name:
        return "yandex_market"

    return "unknown"

def infer_effective_date(filename: str):
    name = filename.lower()

    patterns = [
        r"(20\d{2})[-_.](\d{2})[-_.](\d{2})",
        r"(\d{2})[-_.](\d{2})[-_.](20\d{2})",
    ]

    for p in patterns:
        m = re.search(p, name)
        if not m:
            continue

        groups = m.groups()

        if len(groups[0]) == 4:
            y, mo, d = groups
        else:
            d, mo, y = groups

        try:
            return datetime(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
        except Exception:
            pass

    return ""

def make_doc_hash(marketplace, document_name, section, topic, rule_text, effective_date, source_url):
    base = "|".join([
        marketplace.strip().lower(),
        document_name.strip().lower(),
        section.strip().lower(),
        topic.strip().lower(),
        rule_text.strip().lower(),
        effective_date.strip().lower(),
        source_url.strip().lower(),
    ])
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:24]

def clean_value(v):
    if v is None:
        return ""
    text = str(v).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text

def build_rule_text(headers, values):
    cells = []
    for i, value in enumerate(values):
        value = clean_value(value)
        if not value:
            continue

        header = clean_value(headers[i]) if headers and i < len(headers) else ""
        if header and len(header) <= 80:
            cells.append(f"{header}: {value}")
        else:
            cells.append(value)

    return " | ".join(cells)

def guess_topic(rule_text):
    parts = [p.strip() for p in rule_text.split("|") if p.strip()]
    if not parts:
        return "Строка документа"
    topic = parts[0]
    if len(topic) > 140:
        topic = topic[:140].rstrip() + "..."
    return topic

def insert_doc(marketplace, document_name, section, topic, rule_text, effective_date, source_url):
    if not marketplace:
        marketplace = "unknown"

    if not rule_text or len(rule_text) < 10:
        return 0

    content_hash = make_doc_hash(
        marketplace, document_name, section, topic, rule_text, effective_date, source_url
    )

    _execute("""
        INSERT OR IGNORE INTO rules_documents
        (marketplace, document_name, section, topic, rule_text, effective_date, source_url, content_hash)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        marketplace,
        document_name,
        section,
        topic,
        rule_text,
        effective_date,
        source_url,
        content_hash,
    ))

    return 1

def import_csv(path: Path):
    marketplace = infer_marketplace(path)
    document_name = path.name
    effective_date = infer_effective_date(path.name)
    source_url = str(path)

    encodings = ["utf-8-sig", "cp1251", "utf-8"]
    rows_imported = 0

    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
                reader = csv.reader(f, dialect)
                all_rows = list(reader)
            break
        except Exception:
            all_rows = None

    if not all_rows:
        return 0

    headers = all_rows[0] if all_rows else []

    for idx, row in enumerate(all_rows[1:], start=2):
        rule_text = build_rule_text(headers, row)
        topic = guess_topic(rule_text)
        section = f"CSV row {idx}"
        rows_imported += insert_doc(
            marketplace, document_name, section, topic, rule_text, effective_date, source_url
        )

    return rows_imported


def split_text_chunks(text, max_chars=1800):
    text = re.sub(r"\s+", " ", text or "").strip()
    if not text:
        return []

    chunks = []
    start = 0

    while start < len(text):
        end = min(start + max_chars, len(text))

        if end < len(text):
            cut = text.rfind(". ", start, end)
            if cut > start + 500:
                end = cut + 1

        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)

        start = end

    return chunks


def import_txt(path: Path):
    marketplace = infer_marketplace(path)
    document_name = path.name
    effective_date = infer_effective_date(path.name)
    source_url = str(path)

    encodings = ["utf-8-sig", "cp1251", "utf-8"]
    text = ""

    for enc in encodings:
        try:
            text = path.read_text(encoding=enc)
            break
        except Exception:
            pass

    rows_imported = 0

    for i, chunk in enumerate(split_text_chunks(text), start=1):
        section = f"TXT chunk {i}"
        topic = guess_topic(chunk)
        rows_imported += insert_doc(
            marketplace,
            document_name,
            section,
            topic,
            chunk,
            effective_date,
            source_url,
        )

    return rows_imported


def import_pdf(path: Path):
    if PdfReader is None:
        raise RuntimeError("pypdf is not installed")

    marketplace = infer_marketplace(path)
    document_name = path.name
    effective_date = infer_effective_date(path.name)
    source_url = str(path)

    reader = PdfReader(str(path))
    rows_imported = 0

    for page_num, page in enumerate(reader.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""

        chunks = split_text_chunks(page_text)

        for chunk_index, chunk in enumerate(chunks, start=1):
            section = f"PDF page {page_num} / chunk {chunk_index}"
            topic = guess_topic(chunk)

            rows_imported += insert_doc(
                marketplace,
                document_name,
                section,
                topic,
                chunk,
                effective_date,
                source_url,
            )

    return rows_imported


def import_xlsx(path: Path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed")

    marketplace = infer_marketplace(path)
    document_name = path.name
    effective_date = infer_effective_date(path.name)
    source_url = str(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    rows_imported = 0

    for ws in wb.worksheets:
        section = ws.title
        headers = None
        header_row_index = None

        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [clean_value(v) for v in row]
            non_empty = [v for v in values if v]

            if len(non_empty) < 2:
                continue

            if headers is None:
                headers = values
                header_row_index = row_index
                continue

            rule_text = build_rule_text(headers, values)

            if not rule_text or len(rule_text) < 10:
                continue

            topic = guess_topic(rule_text)
            full_section = f"{section} / row {row_index}"

            rows_imported += insert_doc(
                marketplace,
                document_name,
                full_section,
                topic,
                rule_text,
                effective_date,
                source_url,
            )

    try:
        wb.close()
    except Exception:
        pass

    return rows_imported

def mark_file_imported(path: Path, fp: str, rows_imported: int):
    st = path.stat()
    _execute("""
        INSERT OR IGNORE INTO rules_imported_files
        (file_path, file_size, file_mtime, file_fingerprint, rows_imported)
        VALUES (?, ?, ?, ?, ?)
    """, (
        str(path),
        st.st_size,
        datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
        fp,
        rows_imported,
    ))

def main():
    init_db()
    ensure_tables()

    files = []
    for path in BASE_DIR.rglob("*"):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXT:
            files.append(path)

    print("Files found:", len(files))

    total_imported = 0
    skipped = 0
    failed = 0

    for path in sorted(files):
        imported_before, fp = already_imported(path)

        if imported_before:
            print("SKIP already imported:", path)
            skipped += 1
            continue

        try:
            suffix = path.suffix.lower()

            if suffix == ".csv":
                rows_imported = import_csv(path)
            elif suffix == ".xlsx":
                rows_imported = import_xlsx(path)
            elif suffix == ".pdf":
                rows_imported = import_pdf(path)
            elif suffix == ".txt":
                rows_imported = import_txt(path)
            else:
                rows_imported = 0

            mark_file_imported(path, fp, rows_imported)
            total_imported += rows_imported
            print("OK:", path, "rows:", rows_imported)

        except Exception as e:
            failed += 1
            print("FAILED:", path, repr(e))

    rows = _fetch_all("SELECT COUNT(*) FROM rules_documents")
    total_docs = rows[0][0]

    print()
    print("Bulk import finished")
    print("new rows imported:", total_imported)
    print("files skipped:", skipped)
    print("files failed:", failed)
    print("rules_documents total:", total_docs)

    print()
    print("=== LAST IMPORTED FILES ===")
    rows = _fetch_all("""
        SELECT file_path, rows_imported, imported_at
        FROM rules_imported_files
        ORDER BY id DESC
        LIMIT 10
    """)
    for r in rows:
        print(r)

if __name__ == "__main__":
    main()
